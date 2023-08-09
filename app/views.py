from flask import request, jsonify, send_file
from . import app
from . import getData
import os
import pandas as pd
import openpyxl

endDate = None
startDate = None
pageSize = None
pageNo = None
personPin = None


@app.route('/api/v1/transaction', methods=['GET'])
def getTransaction():
    endDate = request.form.get('endDate')
    startDate = request.form.get('startDate')
    pageSize = request.form.get('pageSize')
    pageNo = request.form.get('pageNo')
    personPin = request.form.get('personPin')

    data = getData.getDataApi(endDate, startDate, pageSize, pageNo, personPin)

    try:
        return jsonify(data), 200
    except Exception as e:
        return jsonify({'message': str(e)}), 500
    

@app.route('/api/v1/transaction/download', methods=['GET'])
def downloadTransaction():
    # endDate = request.form.get('endDate')
    # startDate = request.form.get('startDate')
    # pageSize = request.form.get('pageSize')
    # pageNo = request.form.get('pageNo')
    # personPin = request.form.get('personPin')

    data = getData.getDataApi(endDate, startDate, pageSize, pageNo, personPin)

    try:
        data_points = data["data"]["data"]

        # Create a DataFrame from the data
        df = pd.DataFrame(data_points)

        # fullname
        df['fullname'] = df['name'] + ' ' + df['lastName']

        # Select columns for export
        selected_columns = [
            "id", "fullname", "devName", "doorName", "eventName", "eventTime", "readerName", "pin",
            "areaName", "deptName",
            ]
        df_selected = df[selected_columns]

        # Specify the desired directory and filename
        custom_directory = app.config['FOLDER_EXPORT']  # Replace with the desired directory
        custom_filename = "custom_exported_data.xlsx"

        # Create the full path to the Excel file
        custom_excel_file = os.path.join(custom_directory, custom_filename)
        
        df_selected.to_excel(custom_excel_file, sheet_name="Data", index=False)

        # Load the workbook
        book = openpyxl.load_workbook(custom_excel_file)
        sheet = book.active

        # Add headers with formatting
        header_font = openpyxl.styles.Font(bold=True)
        header_fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        for col_idx, column in enumerate(df_selected.columns, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = column
            cell.font = header_font
            cell.fill = header_fill

        # Auto-adjust column widths
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Get the column name
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        # Save the modified workbook
        book.save(custom_excel_file)

        return send_file (
            custom_excel_file,
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            attachment_filename=custom_filename,
            as_attachment=True
        )
    except Exception as e:
        return jsonify({'message': str(e)}), 500